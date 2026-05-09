"""Створення PNG з діапазону комірок Excel.

Піплайн:
  xlsx -> LibreOffice headless -> PDF (1 сторінка, 200 dpi)
       -> pdf2image -> PIL Image
       -> crop по піксельних координатах діапазону
       -> PNG
Зберігає оригінальні кольори, шрифти, межі та стилі файлу.
"""

import asyncio
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Optional, Tuple

from loguru import logger

_config: dict = {"xlsx_path": None, "sheet": None, "cell_range": None}
_LO_BIN: str = shutil.which("libreoffice") or shutil.which("soffice") or "libreoffice"

# Excel стандартні розміри за замовчуванням
_DEFAULT_COL_WIDTH_PX = 64   # ~8.43 символи
_DEFAULT_ROW_HEIGHT_PX = 20  # 15pt
# Коефіцієнти переведення Excel одиниць → px (при 96 dpi)
_COL_UNIT_TO_PX = 7.5        # 1 символ ~7.5px
_ROW_PT_TO_PX = 96 / 72      # 1pt = 96/72 px


def set_xlsx_config(
    xlsx_path: Optional[str],
    sheet: Optional[str] = None,
    cell_range: Optional[str] = None,
) -> None:
    _config["xlsx_path"] = xlsx_path
    _config["sheet"] = sheet
    _config["cell_range"] = cell_range
    logger.info(
        f"[xlsx_screenshot] config: path={xlsx_path}, "
        f"sheet={sheet}, range={cell_range}"
    )


async def make_schedule_screenshot() -> Optional[str]:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _render_sync)


# ─── Допоміжні функції ─────────────────────────────────────────────

def _col_letter_to_index(col: str) -> int:
    """'A'->1, 'AK'->37"""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _parse_range(cell_range: str) -> Optional[Tuple[int, int, int, int]]:
    """'B4:AK14' -> (col_start=2, row_start=4, col_end=37, row_end=14)"""
    m = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', cell_range.upper())
    if not m:
        return None
    return (
        _col_letter_to_index(m.group(1)), int(m.group(2)),
        _col_letter_to_index(m.group(3)), int(m.group(4)),
    )


def _get_crop_box_px(
    ws, col_start: int, row_start: int, col_end: int, row_end: int, dpi: int
) -> Tuple[int, int, int, int]:
    """
    Розраховує піксельні координати (left, top, right, bottom)
    для crop за розмірами колонок/рядків з openpyxl.
    dpi — роздільна здатність зображення PDF->PNG.
    """
    scale = dpi / 96.0  # коефіцієнт масштабування відносно Excel 96dpi

    def col_width_px(ci: int) -> float:
        """ci — 1-based індекс колонки"""
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(ci)
        cd = ws.column_dimensions.get(col_letter)
        if cd and cd.width:
            return cd.width * _COL_UNIT_TO_PX
        return _DEFAULT_COL_WIDTH_PX

    def row_height_px(ri: int) -> float:
        """ri — 1-based індекс рядка"""
        rd = ws.row_dimensions.get(ri)
        if rd and rd.height:
            return rd.height * _ROW_PT_TO_PX
        return _DEFAULT_ROW_HEIGHT_PX

    # Відступ зліва: сума ширин колонок 1..(col_start-1)
    left_px = sum(col_width_px(ci) for ci in range(1, col_start))
    # Відступ згори: сума висот рядків 1..(row_start-1)
    top_px = sum(row_height_px(ri) for ri in range(1, row_start))
    # Права: + ширини колонок col_start..col_end
    right_px = left_px + sum(col_width_px(ci) for ci in range(col_start, col_end + 1))
    # Низ: + висоти рядків row_start..row_end
    bottom_px = top_px + sum(row_height_px(ri) for ri in range(row_start, row_end + 1))

    return (
        int(left_px * scale),
        int(top_px * scale),
        int(right_px * scale),
        int(bottom_px * scale),
    )


# ─── Головна функція рендерингу ──────────────────────────────────────

def _render_sync() -> Optional[str]:
    """xlsx → PDF (LibreOffice) → PNG (pdf2image) → crop по діапазону."""
    import openpyxl
    from pdf2image import convert_from_path

    xlsx_path = _config.get("xlsx_path")
    sheet_name = _config.get("sheet")
    cell_range = (_config.get("cell_range") or "").strip().upper()

    if not xlsx_path:
        logger.warning("[xlsx_screenshot] xlsx_path not configured")
        return None

    xlsx_file = Path(xlsx_path)
    if not xlsx_file.exists():
        logger.error(f"[xlsx_screenshot] file not found: {xlsx_file}")
        return None

    # Читаємо розміри колонок/рядків до конвертації
    parsed = _parse_range(cell_range) if cell_range else None
    crop_ws = None
    if parsed:
        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            crop_ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        except Exception as e:
            logger.warning(f"[xlsx_screenshot] could not read dimensions: {e}")

    DPI = 200

    with tempfile.TemporaryDirectory() as tmp_str:
        tmp_dir = Path(tmp_str)
        tmp_xlsx = tmp_dir / xlsx_file.name
        shutil.copy2(xlsx_file, tmp_xlsx)

        # LibreOffice: xlsx → PDF (весь аркуш, без обмежень)
        lo_result = subprocess.run(
            [
                _LO_BIN, "--headless", "--norestore", "--nofirststartwizard",
                "--convert-to", "pdf",
                "--outdir", str(tmp_dir),
                str(tmp_xlsx),
            ],
            capture_output=True, text=True, timeout=60,
        )
        if lo_result.returncode != 0:
            logger.error(f"[xlsx_screenshot] LibreOffice error: {lo_result.stderr}")
            return None

        pdf_path = tmp_dir / (tmp_xlsx.stem + ".pdf")
        if not pdf_path.exists():
            logger.error(f"[xlsx_screenshot] PDF not found: {pdf_path}")
            return None

        # PDF → PIL Image
        try:
            pages = convert_from_path(str(pdf_path), dpi=DPI, first_page=1, last_page=1)
        except Exception as e:
            logger.error(f"[xlsx_screenshot] pdf2image error: {e}")
            return None

        if not pages:
            return None

        img = pages[0]
        logger.info(f"[xlsx_screenshot] full page: {img.width}x{img.height}px")

        # Crop по діапазону
        if parsed and crop_ws is not None:
            col_start, row_start, col_end, row_end = parsed
            box = _get_crop_box_px(crop_ws, col_start, row_start, col_end, row_end, DPI)
            logger.info(f"[xlsx_screenshot] crop box px: {box}")

            # Перевіряємо межі
            left, top, right, bottom = box
            right = min(right, img.width)
            bottom = min(bottom, img.height)

            if right > left and bottom > top:
                img = img.crop((left, top, right, bottom))
                logger.info(f"[xlsx_screenshot] cropped: {img.width}x{img.height}px")
            else:
                logger.warning(f"[xlsx_screenshot] invalid crop box {box}, using full page")

        # Зберігаємо PNG
        out = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        out_path = Path(out.name)
        out.close()
        img.save(out_path, "PNG", optimize=True)
        logger.info(f"[xlsx_screenshot] saved -> {out_path}")
        return str(out_path)
