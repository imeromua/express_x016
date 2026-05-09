"""XLSX-парсер табеля графіків.

Структура файлу (реальна, Травень 2026):
    A1  — порожній (іноді 'в')
    C2  — заголовок: "Табель Х016 за Травень 2026 ..."
    Row 4 — заголовки колонок: №, П.І.Б., Посада, Пт/Сб/../Нд, ...
    Row 5 — номери дат: 1, 2, 3 ... 31
    Row 6..N — дані співробітників (до першого рядка де A і B порожні)

Значення клітинок-змін:
    число (9, 12) — робочий день, кількість годин
    'в' / 'В'    — вихідний
    'ТН'          — тимчасова непрацездатність (лікарняний)
    'ЩВ'          — щорічна відпустка
    'Відпустка'   — відпустка
    None          — порожньо (вихідний або не заповнено)
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import List

import openpyxl
from loguru import logger

# Колонки A=0, B=1, C=2, D=3 ... AH=33  (дні місяця 1..31)
_COL_NUM = 0       # №
_COL_PIB = 1       # П.І.Б.
_COL_POS = 2       # Посада
_COL_DAY_START = 3 # D — день 1
_COL_DAY_END = 33  # AH — день 31

_ROW_HEADER = 4    # рядок заголовків (1-based)
_ROW_DATES = 5     # рядок з номерами дат
_ROW_DATA_START = 6  # перший рядок даних

_DAY_NAMES_UK = [
    "Понеділок", "Вівторок", "Середа",
    "Четвер", "П'ятниця", "Субота", "Неділя",
]

# Нормалізація статусів
_OFF_VALUES = frozenset(["в", "В", ""])
_SICK_VALUES = frozenset(["ТН", "тн", "л", "Л"])
_VACATION_VALUES = frozenset(["ЩВ", "щв", "Відпустка", "відпустка", "ВІДПУСТКА"])


class XlsxParseError(Exception):
    """Помилка парсингу файлу."""


def _clean_pib(raw: str) -> str:
    """Прибирає номери карток, ЦВЗ, зайві пробіли з П.І.Б."""
    cleaned = re.sub(r'\s+\d+\s*', ' ', raw)  # видаляємо числа-коди
    cleaned = cleaned.replace('ЦВЗ', '').replace('цвз', '')
    return ' '.join(cleaned.split())  # нормалізуємо пробіли


def _parse_shift(value) -> tuple[str, bool]:
    """
    Повертає (status, is_working).
    status: 'work' | 'off' | 'sick' | 'vacation'
    """
    if value is None:
        return "off", False
    s = str(value).strip()
    if s in _OFF_VALUES:
        return "off", False
    if s in _SICK_VALUES:
        return "sick", False
    if s in _VACATION_VALUES:
        return "vacation", False
    # Спробуємо число — робочий день
    try:
        float(s)
        return "work", True
    except ValueError:
        pass
    # Невідоме значення — логуємо і вважаємо вихідним
    logger.warning(f"[xlsx_parser] Невідоме значення зміни: {value!r} → off")
    return "off", False


def _weekday_name(d: date) -> str:
    return _DAY_NAMES_UK[d.weekday()]


def parse_schedule_xlsx(file_bytes: bytes) -> List[dict]:
    """
    Парсить .xlsx-файл табеля та повертає список dict для UPSERT.

    Кожен dict містить:
        pib, work_date, status, day_name, is_working

    Raises:
        XlsxParseError: якщо файл некоректний або порожній.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise XlsxParseError(f"Не вдалося відкрити файл: {e}") from e

    ws = wb.active

    # ── 1. Визначаємо місяць і рік з заголовка C2 ─────────────────────
    title = ws.cell(row=2, column=3).value or ""
    logger.info(f"[xlsx_parser] Заголовок: {title!r}")

    # ── 2. Читаємо рядок дат (row 5) → маппінг col_idx → day_number ───
    date_map: dict[int, int] = {}  # col_index (0-based) → день місяця
    dates_row = list(ws.iter_rows(min_row=_ROW_DATES, max_row=_ROW_DATES, values_only=True))[0]
    for col_idx in range(_COL_DAY_START, _COL_DAY_END + 1):
        val = dates_row[col_idx]
        if isinstance(val, (int, float)) and 1 <= int(val) <= 31:
            date_map[col_idx] = int(val)

    if not date_map:
        raise XlsxParseError("Не знайдено рядок дат (очікується row 5)")

    # ── 3. Визначаємо рік і місяць з назви листа або заголовка ─────────
    sheet_name = ws.title  # напр. "Травень_2026"
    year, month = _extract_year_month(title, sheet_name)
    logger.info(f"[xlsx_parser] Рік: {year}, Місяць: {month}")

    # ── 4. Парсимо рядки співробітників ────────────────────────────────
    rows: List[dict] = []
    skipped = 0

    for row_num in range(_ROW_DATA_START, ws.max_row + 1):
        row_data = list(ws.iter_rows(
            min_row=row_num, max_row=row_num, values_only=True
        ))[0]

        num_val = row_data[_COL_NUM]
        pib_raw = row_data[_COL_PIB]

        # Стоп-умова: рядок без номера і ПІБ
        if num_val is None and (pib_raw is None or str(pib_raw).strip() == ""):
            break

        # Пропускаємо якщо немає ПІБ
        if not pib_raw or str(pib_raw).strip() in ("", "None", "nan"):
            skipped += 1
            continue

        pib = _clean_pib(str(pib_raw))

        # Парсимо кожен день
        for col_idx, day_num in date_map.items():
            try:
                work_date = date(year, month, day_num)
            except ValueError:
                continue  # 31 лютого і т.п.

            cell_val = row_data[col_idx] if col_idx < len(row_data) else None
            status, is_working = _parse_shift(cell_val)
            day_name = _weekday_name(work_date)

            rows.append({
                "pib": pib,
                "work_date": work_date,
                "status": status,
                "is_working": is_working,
                "day_name": day_name,
            })

    if not rows:
        raise XlsxParseError("Жодного коректного рядка не знайдено")

    logger.info(
        f"[xlsx_parser] Розпізнано {len(rows)} записів "
        f"({len(rows) // max(len(date_map), 1)} співробітників), "
        f"пропущено рядків: {skipped}"
    )
    return rows


# ── Приватні хелпери ──────────────────────────────────────────────────────────

_MONTH_MAP = {
    "січень": 1, "лютий": 2, "березень": 3, "квітень": 4,
    "травень": 5, "червень": 6, "липень": 7, "серпень": 8,
    "вересень": 9, "жовтень": 10, "листопад": 11, "грудень": 12,
    # назви листів типу Травень_2026
    "traven": 5, "lypen": 7, "serpen": 8,
}


def _extract_year_month(title: str, sheet_name: str) -> tuple[int, int]:
    """Витягує (рік, місяць) з заголовка або назви листа."""
    # Шукаємо рік 4-значний
    year_match = re.search(r'(20\d{2})', title + " " + sheet_name)
    year = int(year_match.group(1)) if year_match else date.today().year

    # Шукаємо місяць по назві
    combined = (title + " " + sheet_name).lower()
    for name, num in _MONTH_MAP.items():
        if name in combined:
            return year, num

    # Fallback: поточний місяць
    logger.warning("[xlsx_parser] Не вдалося визначити місяць, використовуємо поточний")
    return year, date.today().month
